"""Deterministic normalization of Jena weather observations."""

from __future__ import annotations

import hashlib
import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

SHEET_NAME = "Jena Climate"
SOURCE_TIME_FORMAT = "%d.%m.%Y %H:%M:%S"
EXPECTED_INTERVAL_MINUTES = 10
MISSING_SENTINELS = {-9999.0, -999.0, 9999.0}
SOURCE_COLUMNS = [
    "Date Time", "p (mbar)", "T (degC)", "Tpot (K)", "Tdew (degC)", "rh (%)",
    "VPmax (mbar)", "VPact (mbar)", "VPdef (mbar)", "sh (g/kg)",
    "H2OC (mmol/mol)", "rho (g/m**3)", "wv (m/s)", "max. wv (m/s)", "wd (deg)",
]
CANONICAL_COLUMNS = [
    "timestamp_utc", "station_id", "latitude", "longitude", "temp_c", "pressure_hpa",
    "humidity_pct", "source", "raw_timestamp_local", "raw_temp_c", "raw_pressure_mbar",
    "raw_humidity_pct", "quality_flags", "source_row",
]
VALID_RANGES = {
    "temp_c": (-60.0, 60.0),
    "pressure_hpa": (800.0, 1100.0),
    "humidity_pct": (0.0, 100.0),
}
PARQUET_SCHEMA = pa.schema([
    pa.field("timestamp_utc", pa.timestamp("us", tz="UTC"), nullable=False),
    pa.field("station_id", pa.string(), nullable=False),
    pa.field("latitude", pa.float64()),
    pa.field("longitude", pa.float64()),
    pa.field("temp_c", pa.float64()),
    pa.field("pressure_hpa", pa.float64()),
    pa.field("humidity_pct", pa.float64()),
    pa.field("source", pa.string(), nullable=False),
    pa.field("raw_timestamp_local", pa.string(), nullable=False),
    pa.field("raw_temp_c", pa.float64()),
    pa.field("raw_pressure_mbar", pa.float64()),
    pa.field("raw_humidity_pct", pa.float64()),
    pa.field("quality_flags", pa.string(), nullable=False),
    pa.field("source_row", pa.int64(), nullable=False),
])


@dataclass(frozen=True)
class Station:
    station_id: str
    latitude: float
    longitude: float
    timezone_name: str


def sha256_file(path: Path) -> str:
    """Return a lowercase SHA-256 digest without loading the file into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_station(path: Path) -> Station:
    """Load and validate the single-station metadata record."""
    frame = pd.read_csv(path)
    if len(frame) != 1:
        raise ValueError("station_metadata.csv must contain exactly one station")
    row = frame.iloc[0]
    required = ("station_id", "latitude", "longitude", "timezone")
    missing = [name for name in required if pd.isna(row.get(name))]
    if missing:
        raise ValueError(f"Station metadata is missing: {', '.join(missing)}")
    return Station(str(row.station_id), float(row.latitude), float(row.longitude), str(row.timezone))


def parse_number(value: Any) -> tuple[float | None, str | None]:
    """Parse a finite sensor value and classify missing or invalid input."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None, "missing" if value is None or str(value).strip() == "" else "non_numeric"
    if not math.isfinite(number):
        return None, "non_finite"
    if number in MISSING_SENTINELS:
        return None, "sentinel"
    return number, None


def to_utc(naive: datetime, zone: ZoneInfo, occurrence: int) -> tuple[datetime, bool]:
    """Resolve repeated DST wall times deterministically and convert to UTC."""
    first = naive.replace(tzinfo=zone, fold=0)
    second = naive.replace(tzinfo=zone, fold=1)
    ambiguous = first.utcoffset() != second.utcoffset()
    localized = second if ambiguous and occurrence % 2 else first
    return localized.astimezone(timezone.utc), ambiguous


def _read_records(workbook_path: Path, station: Station) -> tuple[pd.DataFrame, Counter, Counter, int]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Required sheet {SHEET_NAME!r} was not found")
    rows = workbook[SHEET_NAME].iter_rows(values_only=True)
    header = list(next(rows))
    if header != SOURCE_COLUMNS:
        workbook.close()
        raise ValueError(f"Unexpected workbook schema: {header!r}")

    zone = ZoneInfo(station.timezone_name)
    occurrences: defaultdict[datetime, int] = defaultdict(int)
    missing = Counter({name: 0 for name in VALID_RANGES})
    suspect = Counter({name: 0 for name in VALID_RANGES})
    records: list[dict[str, Any]] = []
    ambiguous_rows = 0

    for source_row, row in enumerate(rows, start=2):
        if all(value is None for value in row):
            continue
        raw_timestamp = row[0]
        local_time = raw_timestamp if isinstance(raw_timestamp, datetime) else datetime.strptime(
            str(raw_timestamp).strip(), SOURCE_TIME_FORMAT
        )
        timestamp_utc, ambiguous = to_utc(local_time, zone, occurrences[local_time])
        occurrences[local_time] += 1
        ambiguous_rows += int(ambiguous)

        pressure, pressure_issue = parse_number(row[1])
        temperature, temperature_issue = parse_number(row[2])
        humidity, humidity_issue = parse_number(row[5])
        values = {
            "temp_c": (temperature, temperature_issue),
            "pressure_hpa": (pressure, pressure_issue),
            "humidity_pct": (humidity, humidity_issue),
        }
        flags: list[str] = []
        for name, (value, issue) in values.items():
            if issue:
                missing[name] += 1
                flags.append(f"{name}:{issue}")
            elif not VALID_RANGES[name][0] <= value <= VALID_RANGES[name][1]:
                suspect[name] += 1
                flags.append(f"{name}:out_of_range")

        records.append({
            "timestamp_utc": timestamp_utc,
            "station_id": station.station_id,
            "latitude": station.latitude,
            "longitude": station.longitude,
            "temp_c": temperature,
            "pressure_hpa": pressure,
            "humidity_pct": humidity,
            "source": "jena_climate_2009_2016",
            "raw_timestamp_local": local_time.isoformat(),
            "raw_temp_c": temperature,
            "raw_pressure_mbar": pressure,
            "raw_humidity_pct": humidity,
            "quality_flags": ";".join(flags),
            "source_row": source_row,
        })
    workbook.close()
    return pd.DataFrame.from_records(records, columns=CANONICAL_COLUMNS), missing, suspect, ambiguous_rows


def normalize_workbook(
    input_path: str | Path,
    output_path: str | Path,
    report_path: str | Path,
    metadata_path: str | Path,
) -> dict[str, Any]:
    """Normalize a workbook, write canonical Parquet, and return its quality report."""
    input_path = Path(input_path)
    output_path = Path(output_path)
    report_path = Path(report_path)
    station = load_station(Path(metadata_path))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    source_hash_before = sha256_file(input_path)

    frame, missing, suspect, ambiguous_rows = _read_records(input_path, station)
    raw_rows = len(frame)
    frame.sort_values(["station_id", "timestamp_utc", "source_row"], kind="mergesort", inplace=True)
    duplicate_mask = frame.duplicated(["station_id", "timestamp_utc"], keep="first")
    duplicate_rows = int(duplicate_mask.sum())
    frame = frame.loc[~duplicate_mask].reset_index(drop=True)

    intervals = frame.groupby("station_id")["timestamp_utc"].diff().dt.total_seconds().div(60)
    gap_intervals = intervals[intervals > EXPECTED_INTERVAL_MINUTES]
    missing_intervals = int(
        ((gap_intervals / EXPECTED_INTERVAL_MINUTES).round() - 1).clip(lower=0).sum()
    )
    table = pa.Table.from_pandas(frame, schema=PARQUET_SCHEMA, preserve_index=False, safe=True)
    pq.write_table(table, output_path, compression="zstd", version="2.6", write_statistics=True)

    source_hash_after = sha256_file(input_path)
    if source_hash_before != source_hash_after:
        raise RuntimeError("Raw workbook changed during normalization")
    report = {
        "source": {
            "path": str(input_path.resolve()),
            "sha256_before": source_hash_before,
            "sha256_after": source_hash_after,
            "unchanged": True,
            "sheet": SHEET_NAME,
        },
        "output": {
            "path": str(output_path.resolve()),
            "sha256": sha256_file(output_path),
            "schema": str(PARQUET_SCHEMA),
            "row_count": len(frame),
        },
        "row_counts": {"raw": raw_rows, "normalized": len(frame), "dropped_duplicate_rows": duplicate_rows},
        "missing_values": dict(missing),
        "duplicates": {"station_timestamp_rows": duplicate_rows, "remaining": 0},
        "timestamp_gaps": {
            "gap_count": len(gap_intervals),
            "estimated_missing_intervals": missing_intervals,
            "expected_interval_minutes": EXPECTED_INTERVAL_MINUTES,
        },
        "suspect_values": dict(suspect),
        "timestamp_checks": {
            "timezone": "UTC",
            "sorted": True,
            "unique_per_station": True,
            "ambiguous_local_rows_resolved": ambiguous_rows,
            "min": frame.timestamp_utc.min().isoformat(),
            "max": frame.timestamp_utc.max().isoformat(),
        },
        "units": {
            "temp_c": "degrees Celsius",
            "pressure_hpa": "hectopascal (1 mbar = 1 hPa)",
            "humidity_pct": "percent relative humidity",
        },
        "normalization": {
            "deterministic": True,
            "duplicate_policy": "stable keep-first after UTC conversion",
            "unit_conversions": {"pressure_mbar_to_hpa": "identity: 1 mbar = 1 hPa"},
        },
    }
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report
